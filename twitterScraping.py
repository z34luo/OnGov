# coding:utf-8
import requests
import re
import time
import json
import datetime
import urllib.request, urllib.parse, urllib.error
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
import itertools
import re



def twitter_flow(USERNAME, PASSWORD, ANALYTICS_ACCOUNT, NUM_DAYS, OUTPUT_DIRECTORY):
    fa = open("twitterLog.txt",'a')
    fa.write('-----------------------------------------------'+'\n')
    fa.write('--------------')
    d = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fa.write(d)
    fa.write('--------------')
    fa.write('\n\n')

    if (NUM_DAYS>90):
        fa.write("Error: NUM_DAYS should less than or equal to 90\n")
        exit()

    user_agent = {
        'User-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.71 Safari/537.36'}
    session = twitter_login(USERNAME, PASSWORD, user_agent,fa)
    start_time, end_time,today,prior = get_date_range(NUM_DAYS)

    #export tweets data
    type = "tweets"
    fa.write('------get tweets data-----\n')
    tweets_data_string = get_analytics_data(type, session, ANALYTICS_ACCOUNT, start_time, end_time, user_agent,fa)
    split_data = format_data(tweets_data_string)
    outfile = get_filename(type, USERNAME,OUTPUT_DIRECTORY, today, prior)

    wb = openpyxl.Workbook()
    ws = wb.active
    pandas_modify(split_data,ws)
    wb.save(outfile)
    fa.write("tweets CSV downloaded as xlsx: "+outfile+'\n\n')


    #export videos data
    type = "videos"
    fa.write('------get videos data-----\n')
    video_data_string = get_analytics_data(type, session, ANALYTICS_ACCOUNT, start_time, end_time, user_agent,fa)

    split_data = format_data(video_data_string)
    outfile = get_filename(type, USERNAME,OUTPUT_DIRECTORY, today, prior)

    wb = openpyxl.Workbook()
    ws = wb.active
    for line in split_data:
        ws.append(line)
    wb.save(outfile)
    fa.write("videos CSV downloaded as xlsx: " + outfile + '\n\n')
    fa.write('\n\n')



def twitter_login(user, pw, user_agent,fa):
    """Start a requests session and login to Twitter with credentials.
    Returned object is logged-in session."""

    tw_url = "https://twitter.com/"
    session = requests.session()
    first_req = session.get(tw_url)

    auth_token_str = re.search(r'<input type="hidden" value="([a-zA-Z0-9]*)" name="authenticity_token"\>',
                               first_req.text)
    authenticity_token = auth_token_str.group(1)

    login_url = 'https://twitter.com/sessions'

    payload = {
        'session[username_or_email]': user,
        'session[password]': pw,
        'remember_me': '1',
        'return_to_ssl': 'true',
        'scribe_log': None,
        'redirect_after_login': '/',
        'authenticity_token': authenticity_token
    }

    login_req = session.post(login_url, data=payload, headers=user_agent)
    fa.write("login_req response: ")
    fa.write(str(login_req.status_code))
    fa.write('\n')


    return session


def get_date_range(num_days):
    """Return date strings in UTC format. The data is returned as
    (start, end)
    with the end date being today and the begin date being 'num_days' prior.
    Twitter's maximum total days is 90."""

    today = datetime.datetime.utcnow()
    prior = today - datetime.timedelta(days=num_days)

    def add_milliseconds(timestamp):  # arbitrary since millisecond precision not necessary
        milli_ts = int(time.mktime(timestamp.timetuple()) * 1000)
        milli_ts = str(milli_ts)
        return milli_ts

    start = add_milliseconds(prior)
    end = add_milliseconds(today)

    today2 = datetime.datetime.now().strftime("%Y%m%d")
    prior2 = prior.strftime("%Y%m%d")
    return (start, end,today2,prior2)


def get_analytics_data(type, session, analytics_account, start_time, end_time, user_agent,fa):
    """Complete the process behind clicking 'Export data' at
    https://analytics.twitter.com/user/USERNAME/tweets   or
    https://analytics.twitter.com/user/USERNAME/videos
    Data is returned as a raw string containing comma-separated data"""

    export_url = "https://analytics.twitter.com/user/" + analytics_account.lower() + "/"+type+"/export.json"
    bundle_url = "https://analytics.twitter.com/user/" + analytics_account.lower() + "/"+type+"/bundle"

    export_data = {
        'start_time': end_time,
        'end_time': start_time,
        'lang': 'en'
    }
    querystring = '?' + urllib.parse.urlencode(export_data)

    status = 'Pending'
    counter = 0
    while status == 'Pending':
        attempt = session.post(export_url + querystring, headers=user_agent)
        try:
            status_dict = json.loads(attempt.text)
            status = status_dict['status']
        except Exception as e:
            fa.write(e + '\n')
            status = 'Error'
            fa.write(str(counter) + str(status) + '\n')
            fa.write('exit line:  status_dict = json.loads(attempt.text)')
            exit()
        counter += 1
        fa.write(str(counter) + str(status) + '\n')
        time.sleep(5)

    if status == 'Error':
        fa.write('exit line:  attempt = session.post(export_url + querystring, headers=user_agent)'+'\n')
        exit()


    fa.write('----finish:session.post(export_url + querystring, headers=user_agent)---- ' + '\n')

    csv_header = {'Content-Type': 'application/csv',
                  'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                  'Accept-Encoding': 'gzip, deflate, sdch',
                  'Accept-Language': 'en-US,en;q=0.8',
                  'Upgrade-Insecure-Requests': '1',
                  'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.71 Safari/537.36'}

    data_req = session.get(bundle_url + querystring, headers=csv_header)
    fa.write('----finish:session.get(bundle_url + querystring, headers=csv_header)-----'+ '\n')
    fa.write("data_req response: "+str(data_req.status_code)+ '\n')

    return data_req.text



def format_data(data_string):
    """Transform raw data string into list-of-lists format"""
    lines = data_string.split('\"\n\"')
    split_data = [re.split(r"\"\s*,\s*\"", line) for line in lines]

    return split_data


def get_filename(type, USERNAME,output_dir, today, prior):
    """Build descriptive filename for xlsx"""
    f_name = type+'_data_' + USERNAME + "_" + prior + '_' + today + '.xlsx'
    full_path = output_dir + '/' + f_name

    return full_path

def pandas_modify(data,ws):
    cols = tuple(data[0])
 #   idx = [i for i in range(1, ws.max_row)]
 #   print(idx)
    data = data[1:]
#    print(data)
    data = (itertools.islice(r, 0, None) for r in data)
#    print(data)
    df1 = pd.DataFrame(data, columns=cols)
    # 倒叙
    df1 = df1.sort_index(axis=0,ascending=False)
#    print(df1)
    # 最后一列最后一行有 "\n 需要去掉
    df1["promoted media engagements"] = df1["promoted media engagements"].str.replace('"\n', '')
#    print(df1["promoted media engagements"])
    # 计算internal engagements
    df1[['likes', 'replies', 'retweets', 'impressions']] = df1[['likes', 'replies', 'retweets', 'impressions']].astype(
        float)
    df1['internal engagements'] = (df1['likes'] + df1['replies'] + df1['retweets']).astype(int)
#    print(df1[['likes', 'replies', 'retweets', 'impressions','internal engagements']])
    # 计算 internal engagements rate
    df1['internal engagements rate'] = (df1['internal engagements'] / df1['impressions']).astype(float)
    # 增加follwer 列
    df = match_followers(df1)
    # 增加 reach percentage 列
    df['Followers'] = df['Followers'].astype(float)
    df['reach percentage'] = (df['impressions'] / df['Followers']).astype(float)
    # 增加 df['category']
    ls = []
    prol = df['promoted impressions']
    for i in range(0, len(prol)):
        if (prol[i] == '-'):
            ls.append('organic')
        else:
            ls.append('promoted')
    df['category'] = ls
    # 增加audience growth 列
    li = []
    for i in range(0, len(df['Followers'])):
        if (i > 0):
            li.append(df['Followers'][i] - df['Followers'][i-1])
        else:
            li.append(df['Followers'][i])
    df['growth followers'] = li
    # 把所有 － 换成nan
    df = df.replace({'promoted impressions': '-',
                     'promoted engagements':'-',
                     'promoted engagement rate':'-',
                     'promoted retweets': '-',
                     'promoted replies': '-',
                     'media views': '-',
                     'promoted likes': '-',
                     'promoted user profile clicks': '-',
                     'promoted url clicks': '-',
                     'promoted hashtag clicks': '-',
                     'promoted detail expands': '-',
                     'promoted permalink clicks': '-',
                     'promoted app opens': '-',
                     'promoted app installs': '-',
                     'promoted follows': '-',
                     'promoted email tweet': '-',
                     'promoted dial phone': '-',
                     'promoted media views': '-',
                     'promoted media engagements': '-'}, np.nan)
    for r in dataframe_to_rows(df,index=True, header=True):
        ws.append(r)

def match_followers(df):
    wbf = openpyxl.load_workbook("ongovFollowers.xlsx")
    wsf = wbf.active
    data = wsf.values
    cols = next(data)[0:]
    data = list(data)
    data = (itertools.islice(r, 0, None) for r in data)
    dff = pd.DataFrame(data, columns=cols)
#    print(dff)

    # df
    timelist = df["time"]
    par = '[0-9]*-[0-9]*-[0-9][0-9]'
    ls = []
    for r in timelist:
        char = re.compile(par).findall(r)
        ls = ls + char
    df['Date'] =ls

    result = pd.merge(df, dff, on='Date')
    return result

if __name__ == '__main__':
    USERNAME = 'uwdatagroup'
    PASSWORD = '#####'
    ANALYTICS_ACCOUNT = USERNAME
    # NUM_DAYS <=90
    NUM_DAYS = 90
    OUTPUT_DIRECTORY = "data"
    twitter_flow(USERNAME, PASSWORD, ANALYTICS_ACCOUNT, NUM_DAYS, OUTPUT_DIRECTORY)

